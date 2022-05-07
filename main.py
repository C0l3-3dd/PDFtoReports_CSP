from calendar import c
from pickle import TRUE
import shutil
import fitz
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Side,Border,Alignment
import database
import os
import logging
import mail
import toPDFV2 as myPdf

def generaExcel(fichero,salida):
    respaldo = "c:\\Pedidos\\Respaldo"
#Abrirmos el documento PDF
#pdf_documento="1.pdf"
    try:
    #   documento = fitz.open(pdf_documento)
        documento = fitz.open(fichero)
        texto =""
        #Recorremos todas las hojas para sacar el texto
        for hoja in  documento:
            texto+= hoja.get_text("text")

        documento.close()

        lista =[]
        respaldoData =[]
        data =str(texto)
        ndata = re.compile('[0-9]{5}\s-\s[0-9]{1,}') #creamos el regex para las tiendas / prodcutos
        data = data.split("\n")#separamos por saltos de linea
        orden = data[0]# scamos el numemero de orden

        #sacamos las tiendas,productos y nuermo de productos
        id_barras = database.Database.getAll(database.Database())
        #print(id_barras)
        for d in data:
            ######  2  #######
            if  d in id_barras or ndata.match(d):
                #print('Hay Coincidencia: ---> ',d)
                lista.append(d)
                
        columnas = [] #columnas para excel con los nombre de los productos
        productos = [] # se llena con las tiendas y el numero de productos
        aux =[]
        #se se paran las columnas y los productos
        for f in lista:
            
            if( not ndata.match(f) ):
                columnas.append(f)
                respaldoData.append(f)
                if( len(aux) != 0 ):
                    productos.append(aux)
                    aux = []
            else:
                aux.append(f)
        
        #se ingresa el ultimo producto a la lista
        productos.append(aux)
        aux = []

        solo_tiendas =[]
        back_prodcutos = productos #respaldamos los prodcutos
        for data in back_prodcutos:
            for d in data:
                temp =d.split(" ") 
                if not temp[0] in solo_tiendas:#si no exitte la tienda en la lista se agrega
                    solo_tiendas.append(temp[0])
                
        solo_tiendas.sort()#ordenamos las tiendas
        noOrden= orden.split(" ")# separamos la orden 

        os.mkdir(os.path.join(salida,noOrden[2]))# creamos carpeta de salida con nombre de la orden
        os.mkdir(os.path.join(respaldo,noOrden[2])) #creamos carpeta de respaldo con el nombre de la orden

        wb = openpyxl.Workbook()#creamos y abrimos el excel
        hoja = wb.active#activamos la hoja
        hoja.title = noOrden[2]#damos nombre a la hoja con el numero de orden


        #primeras columnas 
        hoja.cell(row=1,column=1,value="Caja")
        hoja.cell(row=1,column=2,value="Tienda")
        hoja.cell(row=1,column=3,value="Número de cajas")

        #llenado de las columna tiendas enn excel
        i = 2
        for tienda in solo_tiendas:
            i += 1
            hoja.cell(row=i ,column=1 ,value=i-2)
            hoja.cell(row=i ,column=2 ,value=int(tienda.lstrip("0")) )
            hoja.cell(row=i ,column=3 ,value=1)

        #llenado de las columnas de los producots
        i = 3
        for col in columnas:
            i += 1
            nombres = database.Database.getProdcuto(database.Database(),str(col))
            if( nombres != None):
                hoja.cell(row=1,column=i,value= str(nombres[0]))
                hoja.cell(row=2,column=i,value= str(nombres[1]) if str(nombres[1]) != "None" else str("") )
            else:
                hoja.cell(row=1,column=i,value= str(col))


        aux_data = productos[0]

        #llenado de matriz de tiendas - producto - NoProductos
        j = 3 
        for aux_data in productos:
            j += 1
            for n in aux_data:
                dara = n.split(" ")
                dara[0] = dara[0].lstrip("0")
                i = 2
                
                for row in solo_tiendas:
                    i +=1
                    if(  int(hoja.cell(row=i,column=2).value) == int(dara[0])):
                        hoja.cell(row=i,column=j,value=int(dara[2]))


        hoja.cell(row=1, column= len(columnas)+5, value= "CODIGO")
        hoja.cell(row=2, column= len(columnas)+4, value= "TOTAL")
        ## en estaparte hiba el codigo que traida los codigos
        i = 2
        h = 67
        if(len(columnas) + 3 < 26):
            letra_fin = len(columnas) +67
        else:
            letra_fin = ((len(columnas) + 3) - 26 ) +64

        if(len(columnas) + 3 < 26 ):
            for k in solo_tiendas:
                i += 1
                h += 1
                formula = '=SUM({}{}:{}{})'.format('D',i,chr(letra_fin),i)
                hoja.cell(row = i, column=len(columnas)+4,value=str(formula))

            h = 67
            i = 3 
            for k in columnas:
                i += 1
                h += 1
                ### revisar el cambio de formula en esta parte como en la parte de suma  por tiendas
                formula = '=SUM({}3:{}{})'.format(chr(h),chr(h),len(solo_tiendas)+2)
                hoja.cell(row = len(solo_tiendas)+3 , column=i,value=str(formula))

            #suma global por tiendas(vertical)
            formula = '=SUM({}3:{}{})'.format( chr(letra_fin + 1) , chr(letra_fin + 1) , len(solo_tiendas)+2 )     
            hoja.cell(row= len(solo_tiendas)+3, column= len(columnas)+4, value= str(formula) )

            #suma global por productos(horizontal)
            formula = '=SUM(D{}:{}{})'.format( len(solo_tiendas) +3, chr(letra_fin) , len(solo_tiendas)+3 )     
            hoja.cell(row= len(solo_tiendas)+4, column= len(columnas)+4, value= str(formula) )

        else:
            for k in solo_tiendas:
                i += 1
                h += 1
                formula = '=SUM({}{}:{}{}{})'.format('D',i,'A',chr(letra_fin),i)
                hoja.cell(row = i, column=len(columnas)+4,value=str(formula))

            h = 67
            i = 3 
            for k in range(3,26):
                i += 1
                h += 1
                formula = '=SUM({}3:{}{})'.format(chr(h),chr(h),len(solo_tiendas)+2)
                hoja.cell(row = len(solo_tiendas)+3 , column=i,value=str(formula))
            h = 64
            for k in range(0,(len(columnas)+3)-26):
                i += 1
                h += 1
                formula = '=SUM(A{}3:A{}{})'.format(chr(h),chr(h),len(solo_tiendas)+2)
                hoja.cell(row = len(solo_tiendas)+3 , column=i,value=str(formula))

            #suma global por tiendas(vertical)
            formula = '=SUM(A{}3:A{}{})'.format( chr(letra_fin + 1) , chr(letra_fin + 1) , len(solo_tiendas)+2 )     
            hoja.cell(row= len(solo_tiendas)+3, column= len(columnas)+4, value= str(formula) )

            #suma global por productos(horizontal)
            formula = '=SUM(D{}:A{}{})'.format( len(solo_tiendas) +3, chr(letra_fin) , len(solo_tiendas)+3 )     
            hoja.cell(row= len(solo_tiendas)+4, column= len(columnas)+4, value= str(formula) )

        #Suma de las cajas
        formula = '=SUM(C3:C{})'.format(len(solo_tiendas)+2)
        hoja.cell(row= len(solo_tiendas)+3, column= 3, value= str(formula) )

        #estilos de la tabla
        background = PatternFill("solid",start_color="00FFFF00")#Color a las celdas 'Amarrillo'
        color_fuente = Font(color='000066CC')#Color a la fuente 'azul'
        background_hojaTemp =  PatternFill("solid",start_color="00C0C0C0")#Color a las celdas 'gris'
        borde = Side(border_style="medium")

        for i in range(1,len(solo_tiendas)+3):
            hoja.cell(row=i,column=1).fill = background

        for i in range(4,len(columnas)+3):
            hoja.cell(row=2,column=i).fill = background
            hoja.cell(row=1,column=i).font = Font(b=True)

        hoja.cell(row=1, column=len(columnas)+5).font = color_fuente

        for i in range(3,len(solo_tiendas)+3):
            hoja.cell(row=i,column=len(columnas)+5).fill = background
            hoja.cell(row=i,column=len(columnas)+5).font = color_fuente
            
        for i in range(3,len(columnas)+5):
            hoja.cell(row=len(solo_tiendas)+3,column=i).font = color_fuente

        for i in range(1,len(solo_tiendas)+3):
            for j in range(1,len(columnas)+5):
                hoja.cell(row= i, column=j).border = Border(top=borde, left=borde, right=borde, bottom=borde)

        for i in range(3,len(columnas)+5):
            hoja.cell(row=len(solo_tiendas)+3,column=i).border = Border(top=borde, left=borde, right=borde, bottom=borde)

        #Tablas de secundaria por numero de productos
        for k in range(0,len(columnas)):
            columna1 = []
            columna2 = []
            columna3 = []
            #sacando los datos de la hoja para las subhojas de los productos
            for i in range(0,len(solo_tiendas)):
                if(hoja.cell(row=3+i,column=4+k).value != None):
                    columna1.append( hoja.cell(row=3+i,column=2).value )
                    columna2.append( hoja.cell(row=3+i,column=4+k).value )
                # columna3.append( folio[i] )
            
            # ordenamiendo de los datos por catindad de produdctos de menor a mayor
            for i in range(0,len(columna2)):
                for j in range(0,len(columna2) - 1):
                    if(columna2[j+1] < columna2[j]):
                        # N productos #
                        aux2 = columna2[j]    
                        columna2[j]=columna2[j+1]
                        columna2[j+1]=aux2
                        # Tiendas #
                        aux1 = columna1[j]    
                        columna1[j]=columna1[j+1]
                        columna1[j+1]=aux1
                        # Codigos #
                        # aux3 = columna3[j]    
                        # columna3[j]=columna3[j+1]
                        # columna3[j+1]=aux3

            hojaTemp = wb.create_sheet(respaldoData[k]) #se crea la hoja nueva 
            #titulos
            hojaTemp["A1"] = "PEDIDO:"  
            hojaTemp["B1"] = int(noOrden[2])
            hojaTemp["D1"] = "SKU:"
            hojaTemp["E1"] = int(respaldoData[k])
            #titulos de tabla
            hojaTemp["B1"].border = Border(bottom=borde)
            hojaTemp["E1"].border = Border(bottom=borde)

            hojaTemp["A3"] = "ALMACE"
            hojaTemp["B3"] = "CANTIDAD PEDIDA"
            hojaTemp["C3"] = "CANT.DE CONTE"
            hojaTemp["D3"] = "CANT.POR CONT."
            hojaTemp["E3"] = "CONTENEDOR"

            hojaTemp["A3"].fill = background_hojaTemp #estilo de color gris en fondo
            hojaTemp["B3"].fill = background_hojaTemp 
            hojaTemp["C3"].fill = background_hojaTemp 
            hojaTemp["D3"].fill = background_hojaTemp 
            hojaTemp["E3"].fill = background_hojaTemp 
            hojaTemp["A3"].font = Font(b=True)#fuente bold
            hojaTemp["B3"].font = Font(b=True)
            hojaTemp["C3"].font = Font(b=True)
            hojaTemp["D3"].font = Font(b=True)
            hojaTemp["E3"].font = Font(b=True)
            hojaTemp.column_dimensions["A"].width =10 #tamaño de la celda
            hojaTemp.column_dimensions["B"].width =20
            hojaTemp.column_dimensions["C"].width =20
            hojaTemp.column_dimensions["D"].width =20
            hojaTemp.column_dimensions["E"].width =15
            hojaTemp["A3"].alignment = Alignment (horizontal = 'center', vertical = 'center') #alineado texto de la celda 
            hojaTemp["B3"].alignment = Alignment (horizontal = 'center', vertical = 'center')
            hojaTemp["C3"].alignment = Alignment (horizontal = 'center', vertical = 'center')
            hojaTemp["D3"].alignment = Alignment (horizontal = 'center', vertical = 'center')
            hojaTemp["E3"].alignment = Alignment (horizontal = 'center', vertical = 'center')

            #llenando tabla con los datos A=tienda, B = prodcutos, c= 1 D = B, E = codigo
            for i in range(0,len(columna1)):
                hojaTemp["A"+ str(i+4)] = int(columna1[i])
                hojaTemp["B"+ str(i+4)] = int(columna2[i])
                hojaTemp["C"+ str(i+4)] = 1
                hojaTemp["D"+ str(i+4)] = int(columna2[i])
            #hojaTemp["E"+ str(i+4)] = int(columna3[i])
                hojaTemp["A"+ str(i+4)].alignment = Alignment (horizontal = 'center', vertical = 'center') #alineado texto de la celda 
                hojaTemp["B"+ str(i+4)].alignment = Alignment (horizontal = 'center', vertical = 'center') 
                hojaTemp["C"+ str(i+4)].alignment = Alignment (horizontal = 'center', vertical = 'center') 
                hojaTemp["D"+ str(i+4)].alignment = Alignment (horizontal = 'center', vertical = 'center') 
                #hojaTemp["E"+ str(i+4)].alignment = Alignment (horizontal = 'center', vertical = 'center') 

            for i in range(0,len(columna3)):#estilos para la columna de codigo
                hojaTemp.cell(row=4+i , column=5).fill = background
                hojaTemp.cell(row=4+i , column=5).font = color_fuente

            for i in range(1,6):#bordes de la tabla
                for j in range(0,len(columna1)+1):
                    hojaTemp.cell(row=3+j,column=i).border = Border(top=borde, left=borde, right=borde, bottom=borde)
                    
            columna1.clear
            columna2.clear
            columna3.clear


    ###### Inicio para los datos de PDF##############
        #print(hoja.max_column, hoja.max_row,"\n")
        titulos =[]
        dataEtiquetas=[]
        maru=[]
        male=[]
        yola=[]
        martha=[]
        consulta = ""
        for i in range(4, hoja.max_column-2):
            #consulta = False
            for j in range( 1,hoja.max_row-1):
                contenido = hoja.cell(row= j, column= i).value
                if contenido != None:
                    if(j <= 2):
                        titulos.append(contenido)     
                        #consulta = True
                        #print(contenido)

                    # if(i == 9):
                    #     print(i,j)

                    if len(titulos) == 2 : 
                        consulta = database.Database.getXname(database.Database(),str(titulos[0]),str(titulos[1]) )
                        nameProduct = str(titulos[0])+"   "+str(titulos[1])
                        titulos.clear()
                        if consulta == None:
                            break
                        
                    if(j >=3):
                        if consulta == 'Maru':
                            maru.append({'tienda':hoja.cell(row= j, column= 2).value,'produtos':contenido,'nameProducto':nameProduct  })
                        # print(hoja.cell(row= j, column= 2).value, contenido )
                        elif consulta == 'Male':
                            male.append({'tienda':hoja.cell(row= j, column= 2).value,'produtos':contenido,'nameProducto':nameProduct  })
                        elif consulta == 'Yola':
                            yola.append({'tienda':hoja.cell(row= j, column= 2).value,'produtos':contenido,'nameProducto':nameProduct  })
                        elif consulta == 'Martha':
                            martha.append({'tienda':hoja.cell(row= j, column= 2).value,'produtos':contenido,'nameProducto':nameProduct  })

                    
                    #consulta = False

        
        # for contenido in maru:
        #     print(contenido)
        # for contenido in male:
        #     print(contenido)
        # for contenido in yola:
        #     print(contenido)
        # for contenido in martha:
        #     print(contenido)

        myPdf.create([maru,male,martha,yola], os.path.join(salida,noOrden[2],noOrden[2]+'_Etiqueta_3x3x3.pdf'))
        
    
    ###### Fin para los datos de PDF##############

        #Guarda el excel
        
        wb.save(os.path.join(salida,noOrden[2],noOrden[2]+'.xlsx')) #salvamos el excel en la salida dentro de la carpeta con su orden
        #wb.save(os.path.join(noOrden[2]+'.xlsx')) #salvamos el excel en la salida dentro de la carpeta con su orden
        shutil.copy( os.path.join(salida,noOrden[2],noOrden[2]+'.xlsx'), os.path.join(respaldo,noOrden[2],noOrden[2]+'.xlsx' ) )#copiamos el excel a reslpado
        shutil.copy( os.path.join(salida,noOrden[2],noOrden[2]+'_Etiqueta_3x3x3.pdf'), os.path.join(respaldo,noOrden[2],noOrden[2]+'_Etiqueta_3x3x3.pdf' ) )#copiamos el pdf de etiqueta a reslpado
        shutil.move(fichero, os.path.join(respaldo,noOrden[2])) #movemo el pdf de entrada a respaldo
        #os.remove(fichero)
    except (Exception ,FileNotFoundError, IndexError, OverflowError, TypeError ) as e:
        logging.exception('EXCEPTION GENERADA: ' +str(e))
        mail.senMail("Ocurrio un error al generar el Excel, revisar lo logs"+ str(e))
    else:
        logging.info('Excel generado correctamente')
        mail.senMail("Excel generado correctamente, orden: "+noOrden[2])